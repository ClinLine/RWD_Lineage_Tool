import pandas as pd
import openpyxl
import networkx as nx  
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from matplotlib.widgets import Button

def load_data(workbook):
    wb = openpyxl.load_workbook(workbook)
    mapping_df = wb['MappingConcepts']
    variables_df = wb['Variables']
    datasets_df = wb['Dataset']    
    link_df = wb['DatasetLink']
    return mapping_df, variables_df, datasets_df, link_df

def Column_names(sheet, headerRow):
    ColNames = []
    ColNames.append("") #add empty value to align names with sheet column names
    for i in range(1, sheet.max_column + 1):
        ColName = sheet.cell(row=headerRow, column=i).value
        ColNames.append(ColName)
   
    return ColNames    

def define_levels(mapping_df, variables_df):
    levels = {}
    dependencies = {}    
    edges = []
    mappingNames = {}
    map_col = Column_names(mapping_df,2)
    var_col = Column_names(variables_df,1)
    ds_col = Column_names(variables_df,1)
    for x in map_col:
        if x == 'Required': ReqCol = map_col.index(x)  
        if x == 'sourceVariables.id': SourceIdVarCol = map_col.index(x)
        if x == 'targetVariable.id': TargetIdVarCol = map_col.index(x)
    for x in var_col:
        if x == 'id': VarIdCol = var_col.index(x) 
        if x == 'parent.name': DsNameCol = var_col.index(x)
    for x in ds_col:
        if x == 'id': DsIdCol = ds_col.index(x)

    for i in range(3,mapping_df.max_row + 1) : # first get all concept names and ids
        concept = mapping_df.cell(row=i, column=1).value # id will be on first column
        conceptName = mapping_df.cell(row=i, column=2).value # id will be on first column     
        if conceptName =="": conceptName=concept   
        mappingNames[concept]=conceptName
    
    for i in range(3,mapping_df.max_row + 1) : # Then check which concepts are required
        concept = mapping_df.cell(row=i, column=1).value # id will be on first column
        conceptName = mapping_df.cell(row=i, column=2).value # id will be on first column        
        if conceptName =="": conceptName=concept
        required = mapping_df.cell(row=i, column=ReqCol).value
        
        if pd.notna(required) and required != '':
            required_concepts = [mappingNames[req.strip()] for req in required.split(",") if req.strip()
            if req.strip() and req.strip().lower() != concept.lower()]
            dependencies[conceptName] = required_concepts
        else:
            if conceptName != '' and conceptName != None:
                dependencies[conceptName] = []
    
    levels = {node: -1 for node in dependencies}    
    
    for concept, required_concepts in dependencies.items():
        if not required_concepts:
            levels[concept] = 1

    while True:
        updated = False
        
        for concept, required_concepts in dependencies.items():
            if levels[concept] == -1:
                if all(levels[req] != -1 for req in required_concepts):
                    levels[concept] = max(levels[req] for req in required_concepts) + 2
                    updated = True
        if not updated:
            break
    
        
    for i in range(3,mapping_df.max_row + 1) :   # add in and output datasets     
        mapping_concept = mapping_df.cell(row=i, column=2).value # id will be on first column
        source_variable = mapping_df.cell(row=i, column=SourceIdVarCol).value
        target_variable = mapping_df.cell(row=i, column=TargetIdVarCol).value
        for j in range(2,variables_df.max_row + 1) :
            if source_variable== variables_df.cell(row=j, column=VarIdCol).value :
                source_dataset = variables_df.cell(row=j, column=DsNameCol).value
                if source_dataset != '' and source_dataset != None and levels.get(source_dataset, 999) > levels.get(mapping_concept, -1) - 1:
                    levels[source_dataset] = levels.get(mapping_concept, -1) - 1
                if source_dataset != '' and source_dataset != None and (source_dataset,mapping_concept) not in edges: edges.append((source_dataset,mapping_concept))

            if target_variable== variables_df.cell(row=j, column=VarIdCol).value :
                target_dataset = variables_df.cell(row=j, column=DsNameCol).value
                if target_dataset != '' and target_dataset != None and levels.get(target_dataset, 999) > levels.get(mapping_concept, -1) - 1:
                    levels[target_dataset] = levels.get(mapping_concept, -1) +1
                if target_dataset != '' and target_dataset != None and (mapping_concept,target_dataset) not in edges: edges.append((mapping_concept,target_dataset))
    return levels, edges

def load_ConceptData(conceptName, mapping_df, variables_df, link_df):
    
    class MyVarArr1:
        transType = {}
        transCd = {}
        transTxt = {}
        conditionCd = {}
        conditionTxt = {}
        sourceVar = {}
        sourceVarLabel = {}
        targetVar = {}
        targetVarLabel = {}
        sourceDs = {}
        targetDs = {}

    InputDs = set()
    class MyLinkArr:
        ds1 = {}
        ds2 = {}
        byvar1 = {}
        byvar2 = {}
        joinType = {}
        label = {}
        descr = {}
    
    map_col = Column_names(mapping_df,2)
    var_col = Column_names(variables_df,1)
    lnk_col = Column_names(link_df,1)
    for x in map_col:
        if x == 'transformationType': transTypeCol  = map_col.index(x)
        if x == 'customTransformationText' : transTxtCol = map_col.index(x)
        if x == 'customTransformationCode' : transCdCol = map_col.index(x)
        if x == 'conditionText' : CondCol = map_col.index(x)
        if x == 'conditionCode' : CondCdCol = map_col.index(x)
        if x == 'sourceVariables.id': SourceIdVarCol = map_col.index(x)
        if x == 'targetVariable.id': TargetIdVarCol = map_col.index(x)
    for x in var_col:
        if x == 'id': VarIdCol = var_col.index(x) 
        if x == 'parent.name': DsNameCol = var_col.index(x)
        if x == 'name': VarNameCol = var_col.index(x)
        if x == 'label': VarLabelCol = var_col.index(x)
    for x in lnk_col:
        if x == 'dataset1': Ds1Col = lnk_col.index(x)
        if x == 'dataset2': Ds2Col = lnk_col.index(x)        
        if x == 'label': LabelCol = lnk_col.index(x)
        if x == 'description': DescCol = lnk_col.index(x)
        if x == 'byvar1': By1Col = lnk_col.index(x)
        if x == 'byvar2': By2Col = lnk_col.index(x)        
        if x == 'joinType': JoinCol = lnk_col.index(x)      
    x=0
    for i in range(3,mapping_df.max_row + 1) :   # add in and output datasets     
        if mapping_df.cell(row=i, column=2).value == conceptName:
            x=x+1
            MyVarArr1.transType[x]=mapping_df.cell(row=i, column=transTypeCol).value
            MyVarArr1.transType[x].replace('UNIQUE_NUMBER', 'UNIQUE NUMBER')
            MyVarArr1.transCd[x]=mapping_df.cell(row=i, column=transCdCol).value
            MyVarArr1.transTxt[x]=mapping_df.cell(row=i, column=transTxtCol).value
            MyVarArr1.conditionCd[x]=mapping_df.cell(row=i, column=CondCol).value
            MyVarArr1.conditionTxt[x]=mapping_df.cell(row=i, column=CondCdCol).value
            for j in range(2,variables_df.max_row + 1) :
                if mapping_df.cell(row=i, column=SourceIdVarCol).value== variables_df.cell(row=j, column=VarIdCol).value :
                    MyVarArr1.sourceVar[x]=variables_df.cell(row=j, column=VarNameCol).value
                    MyVarArr1.sourceVarLabel[x]=variables_df.cell(row=j, column=VarLabelCol).value
                    MyVarArr1.sourceDs[x] = variables_df.cell(row=j, column=DsNameCol).value
                    if variables_df.cell(row=j, column=DsNameCol).value != None: InputDs.add(variables_df.cell(row=j, column=DsNameCol).value)
                if mapping_df.cell(row=i, column=TargetIdVarCol).value== variables_df.cell(row=j, column=VarIdCol).value :
                    MyVarArr1.targetVar[x]=variables_df.cell(row=j, column=VarNameCol).value
                    MyVarArr1.targetVarLabel[x]=variables_df.cell(row=j, column=VarLabelCol).value
                    MyVarArr1.targetDs[x] = variables_df.cell(row=j, column=DsNameCol).value     

    x=0
    
    for i in range(2,link_df.max_row + 1) : 
        if link_df.cell(row=i, column=Ds1Col).value in InputDs and link_df.cell(row=i, column=Ds2Col).value in InputDs:
            x=x+1
            MyLinkArr.ds1[x]=link_df.cell(row=i, column=Ds1Col).value
            MyLinkArr.ds2[x]=link_df.cell(row=i, column=Ds2Col).value
            MyLinkArr.byvar1[x]=link_df.cell(row=i, column=By1Col).value
            MyLinkArr.byvar2[x]=link_df.cell(row=i, column=By2Col).value
            MyLinkArr.joinType[x]=link_df.cell(row=i, column=JoinCol).value            
            MyLinkArr.label[x]=link_df.cell(row=i, column=LabelCol).value
            MyLinkArr.descr[x]=link_df.cell(row=i, column=DescCol).value
        

    return MyVarArr1, MyLinkArr

def build_main_graph(mapping_df, variables_df, startTxt):
    G = nx.DiGraph()  # Create a directed graph
    levels, edges = define_levels(mapping_df, variables_df)
    
    G.add_node(startTxt,type="Start",color="white")
    for level in levels:
        if levels[level] / 2 == int(levels[level] / 2):
            G.add_node(level, type="Dataset", color="lightgreen")
        else:
            G.add_node(level, type="MappingConcept", color="lightblue")

    # Add edges to the graph
    for edge in edges:
    # Remove edges where both nodes are MappingConcepts
      edges = [edge for edge in edges if not (G.nodes[edge[0]].get('type') == 'MappingConcept' and G.nodes[edge[1]].get('type') == 'MappingConcept')]
      G.add_edges_from(edges)
   
    return G, levels

def draw_graph_with_buttons(G, levels, startTxt, mapping_df, variables_df, link_df):
    fig, ax = plt.subplots(figsize=(12, 8))
    plt.subplots_adjust(bottom=0.2)
    
    level_nodes = {}
    level_nodes[0]=[]
    level_nodes[0].append(startTxt)
    for node, level in levels.items():
        if level not in level_nodes:
            level_nodes[level] = []
        level_nodes[level].append(node)
    
    pos = {}
    
    max_nodesDs=1
    max_nodesMC=1
    for level, nodes_at_level in sorted(level_nodes.items()):
        if level / 2 == int(level / 2):
            if len(nodes_at_level)>max_nodesDs: max_nodesDs=len(nodes_at_level)
        else:
            if len(nodes_at_level)>max_nodesMC: max_nodesMC=len(nodes_at_level)
    if max_nodesDs>max_nodesMC: max_nodesMC=max_nodesDs
    y_spaceDs=10*(1/max_nodesDs)

    for level, nodes_at_level in sorted(level_nodes.items()):
        for i, node in enumerate(nodes_at_level):
            if level==0: #first column
                if i==0: pos[node] = (level,0)
                else: pos[node] = (level,( -i * y_spaceDs)+1)
            elif level / 2 == int(level / 2): # all other datasets
                pos[node] = (level,( (i+1) * -y_spaceDs)+1)
            else:  # and the nodes
                pos[node] = (level,( i * -y_spaceDs)-1)
      
    mapping_concepts = [node for node in G.nodes if G.nodes[node].get('type') == "MappingConcept"]
    datasets = [node for node in G.nodes if G.nodes[node].get('type') == "Dataset"]
    starts =  [node for node in G.nodes if G.nodes[node].get('type') == "Start"]

    nx.draw_networkx_nodes(G, pos, nodelist=mapping_concepts, node_color='#2FAC66', node_size=3500, node_shape="o", alpha=0.8, ax=ax)  # squares
    nx.draw_networkx_nodes(G, pos, nodelist=datasets, node_color='#66B2E4', node_size=3500, node_shape="s", alpha=0.8, ax=ax)  # Squares
    nx.draw_networkx_nodes(G, pos, nodelist=starts, node_color='white', node_size=200, node_shape="s", alpha=0.8, ax=ax)  # Squares


    nx.draw_networkx_labels(G, pos, font_size=8, font_color='black', font_weight='bold', ax=ax)
    nx.draw_networkx_edges(G, pos, edge_color='gray', arrowstyle='-|>', node_size=2000, node_shape="s", arrowsize=10, connectionstyle="arc3,rad=0.2", ax=ax)
       
    ax.set_title("Mapping Concepts and Datasets (Click a mapping node for more details)")
    fig.canvas.mpl_connect('button_press_event', lambda event: on_click(event, G, pos, mapping_df, variables_df, link_df))
    plt.show()

def on_click(event, G, pos, mapping_df, variables_df, link_df):
    """Handles node clicks and triggers the detailed graph for MappingConcepts."""
    #print(f"Click detected at ({event.xdata}, {event.ydata})")  # Debugging line
    
    if event.xdata is None or event.ydata is None:
        return  # Ignore clicks outside the graph

    for node, (x, y) in pos.items():
        # print(f"Checking node '{node}' at ({x}, {y})")  # Debugging line
        
        if abs(event.xdata - x) < 0.3 and abs(event.ydata - y) < 0.3:
            node_type = G.nodes[node].get('type', None)
            
            if node_type == "MappingConcept":
               # print(f"Clicked Mapping Concept: {node}")  # Debugging line
                MapConceptVars, MapLinkVars = load_ConceptData(node, mapping_df, variables_df, link_df)
                #run_graph(node)  # Call detailed graph function in link1.py
                G2, pos, CntRows = build_level2_graph(MapConceptVars,MapLinkVars)
                draw_level2_graph(G2, node, pos, CntRows)
            break

def build_level2_graph(MapConceptVars,MapLinkVars):
    """Constructs a directed graph based on a single MappingConcept ID."""
    G = nx.DiGraph()
    pos = {}
    y_cursor = 0
    SourceY={}
    TargetY={}
    DsIn = []
    DsOut = []
    dataset_nodes = {}  # Track added dataset nodes to avoid duplication
    x_offset = {
        'dataset_source': -2, 'variable_source': -1, 'transformation': 0,
        'variable_target': 1, 'dataset_target': 2, 'link': -3
    }
    var_offset = np.array([[-1,0,1,0]]) # source no, Number of mappings, offset for mapping, counter

    # define positions for datasets in grid
    CntRows = len(MapConceptVars.sourceVar) # Total number of rows mappings.
    for i in range(0,len(MapConceptVars.sourceVar)):
        x=i+1
        if MapConceptVars.sourceDs[x] != None and MapConceptVars.sourceDs[x] not in DsIn:
            DsIn.append(MapConceptVars.sourceDs[x])
        if MapConceptVars.targetDs[x] != None and MapConceptVars.targetDs[x] not in DsOut:
            DsOut.append(MapConceptVars.targetDs[x])

    for i in range(0,len(DsIn)):
        var_offset=np.append(var_offset,[[i,0,1,0]],axis=0)
    
    for s in range(0,len(MapConceptVars.sourceVar)):
        dsId=-1
        x=s+1
        if MapConceptVars.sourceDs[x] != None:
            for i in range(0,len(DsIn)):
                if MapConceptVars.sourceDs[x]==DsIn[i]: 
                    var_offset[i+1,1] = var_offset[i+1,1]+1
        else:
            var_offset[0,1] = var_offset[0,1]+1
    for s in range(1, len(var_offset)):
        var_offset[s,2] = var_offset[s-1,1]+var_offset[s-1,2]
    # print(var_offset)

    def GetPosVar(sourceDs):
        if sourceDs != None:
            for i in range(0,len(DsIn)):
                if sourceDs==DsIn[i]:
                    pos=var_offset[i+1,2]+var_offset[i+1,3]
                    var_offset[i+1,3]=var_offset[i+1,3]+1
        else: 
            pos=var_offset[0,2]+var_offset[0,3]
            var_offset[0,3]=var_offset[0,3]+1
        return pos
    
    def GetPosInDs(sourceDs):
        if sourceDs != None:
            for i in range(0,len(DsIn)):
                if sourceDs==DsIn[i]:
                    pos=(var_offset[i+1,1]/4)+var_offset[i+1,2]
        else: 
            pos=var_offset[0,1]/2
        return pos
    
    # define order for variables and transformations in grid.
        
    for i in range(0, len(MapConceptVars.sourceVar)): 
        x=i+1
        y_cursor=y_cursor+1
        source_var_name = MapConceptVars.sourceVar[x]
        #source_label = MapConceptVars.sourceVarLabel[x]
        target_var_name = MapConceptVars.targetVar[x]
        target_var_id = "t_" + target_var_name
        #target_label = MapConceptVars.targetVarLabel[x]
        transformation_type = MapConceptVars.transType[x]

        # Define dataset labels
        source_dataset_id = MapConceptVars.sourceDs[x]
        target_dataset_id = MapConceptVars.targetDs[x]
        XPos=GetPosVar(source_dataset_id)
        InDsPos=GetPosInDs(source_dataset_id)-1
        
        #source_dataset_label = parent_id_to_label.get(source_dataset_id, source_dataset_id)
        #target_dataset_label = parent_id_to_label.get(target_dataset_id, target_dataset_id) if target_dataset_id else None

        # Mark source datasets for link connections
        if source_dataset_id:
            if source_dataset_id not in dataset_nodes:
                G.add_node(source_dataset_id, type='dataset')
                dataset_nodes[source_dataset_id] = True
                pos[source_dataset_id] = (x_offset['dataset_source'],-InDsPos)

        if target_dataset_id and target_dataset_id not in dataset_nodes:
            G.add_node(target_dataset_id, type='dataset')
            dataset_nodes[target_dataset_id] = True
            pos[target_dataset_id] = (x_offset['dataset_target'],-(CntRows/2))

        # Add transformation node + edge
        G.add_node(x, type='transformation', label=transformation_type)
        pos[x] = (x_offset['transformation'], -XPos)

        # Add source variable node + edge
        if source_var_name:
            G.add_node(source_var_name, type='svariable')
            pos[source_var_name] = (x_offset['variable_source'], -XPos)
            G.add_edge(source_var_name, x)

            if source_dataset_id:
                G.add_edge(source_dataset_id, source_var_name)
                
            
        # Add target variable node + edge
        if target_var_name:
            G.add_node(target_var_id, type='tvariable',label=target_var_name)            
            pos[target_var_id] = (x_offset['variable_target'], -XPos)
            if target_dataset_id:
                G.add_edge(target_var_id, target_dataset_id)
            
            # Add edge
            G.add_edge(x, target_var_id)
    
    # Add dataset link nodes
    for i in range(0, len(MapLinkVars.ds1)): 
        x=i+1
        print(MapLinkVars.ds1[x])
        print(MapLinkVars.ds2[x])

        G.add_node(MapLinkVars.label[x], type='link', label=MapLinkVars.label[x])
        G.add_edge(MapLinkVars.ds1[x], MapLinkVars.label[x])
        G.add_edge(MapLinkVars.ds2[x], MapLinkVars.label[x])
        pos[MapLinkVars.label[x]]=(x_offset['dataset_source'], -(GetPosInDs(MapLinkVars.ds1[x])+GetPosInDs(MapLinkVars.ds2[x]))/2)
    #        
    #     print (link_df)
    #     print("Checking link:", dataset1_label, dataset2_label)
    #     print("Available dataset nodes:", dataset_nodes)
    #     print(f"Row: {dataset1_id}, {dataset2_id}, linkid: {linkid}")
    #     print(f"Labels: {dataset1_label}, {dataset2_label}")
    #     print(f"Exists in dataset_nodes? {dataset1_label in dataset_nodes}, {dataset2_label in dataset_nodes}")
       # print (dataset_nodes)
       
    return G, pos, CntRows


def draw_level2_graph(G, mapping_concept_id, pos, CntRows):
    """Draws a clean, grouped, left-to-right graph for a given MappingConceptID."""

    if G is None:
        return

    # variable_to_label = dict(zip(variables_df['id'], variables_df['label']))
    # variable_to_dataset = dict(zip(variables_df['id'], variables_df['parent.id']))
    # dataset_id_to_label = dict(zip(datasets_df['id'], datasets_df['label']))



    # mc_df = mapping_df[mapping_df['MappingConceptID'] == mapping_concept_id].copy()
    # mc_df['sourceDataset'] = mc_df['sourceVariables.id'].map(variable_to_dataset)
    # mc_df = mc_df.sort_values(by=['sourceDataset', 'sourceVariables.id'])

    # for _, row in mc_df.iterrows():
    #     svar = row['sourceVariables.id']
    #     tvar = row['targetVariable.id'] if pd.notna(row['targetVariable.id']) else None
    #     trans_type = row['transformationType'] if pd.notna(row['transformationType']) else 'NONE'

    #     source_dataset_id = variable_to_dataset.get(svar)
    #     source_dataset_label = dataset_id_to_label.get(source_dataset_id, source_dataset_id)

    #     target_dataset_id = variable_to_dataset.get(tvar) if tvar else None
    #     target_dataset_label = dataset_id_to_label.get(target_dataset_id, target_dataset_id) if target_dataset_id else None

    #     svar_label = variable_to_label.get(svar)
    #     if pd.isna(svar_label):
    #         continue
    #     svar_node = f"{svar_label}_source"

    #     tvar_node = None
    #     if tvar:
    #         tvar_label = variable_to_label.get(tvar)
    #         if pd.notna(tvar_label):
    #             tvar_node = f"{tvar_label}_target"

    #     transformation_node = f"{trans_type}_{svar}_{tvar if tvar else 'NA'}"

    #     if source_dataset_label:
    #         pos[source_dataset_label] = (x_offset['dataset_source'], -y_cursor)
    #         source_dataset_y_positions[source_dataset_label] = -y_cursor
    #     pos[svar_node] = (x_offset['variable_source'], -y_cursor)
    #     pos[transformation_node] = (x_offset['transformation'], -y_cursor)
    #     if tvar_node:
    #         pos[tvar_node] = (x_offset['variable_target'], -y_cursor)
    #     if target_dataset_label:
    #         pos[target_dataset_label] = (x_offset['dataset_target'], -y_cursor)

    #     y_cursor += 2
    
    # for node in G.nodes:
    #     if G.nodes[node].get('type') == 'link':
    #         print (node)
    #         linked_datasets = [n for n in G.predecessors(node) if G.nodes[n].get('type') == 'dataset']
    #         y_vals = [source_dataset_y_positions.get(ds) for ds in linked_datasets if ds in source_dataset_y_positions]
    #         if y_vals:
    #             pos[node] = (x_offset['link'], sum(y_vals) / len(y_vals))

    fig, ax = plt.subplots(figsize=(15, 8))

    safe_edges = [(u, v) for u, v in G.edges() if u in pos and v in pos]

    def side_offset(node, direction):
        ntype = G.nodes[node].get('type')
        if ntype in ['svariable', 'tvariable']:
            return 0.35 if direction == 'out' else -0.35
        elif ntype == 'transformation':
            return 0.25 if direction == 'out' else -0.25
        elif ntype == 'dataset':
            return 0.5 if direction == 'out' else -0.5
        else:
            return 0

    for u, v in safe_edges:
        x1, y1 = pos[u]
        x2, y2 = pos[v]

        u_type = G.nodes[u].get('type')
        v_type = G.nodes[v].get('type')

        # Apply offsets only if not a dataset
        offset_u = 0.2 if u_type == 'dataset' else side_offset(u, 'out')
        offset_v = -0.2 if v_type == 'dataset' else side_offset(v, 'in')

        start = (x1 + offset_u, y1)
        end = (x2 + offset_v, y2)

        ax.annotate("",
            xy=end, xytext=start,
            arrowprops=dict(arrowstyle="->", lw=0.8, color="black"),
        )


    def draw_nodes(nodelist, color, shape, size, alpha):
        safe_nodes = [n for n in nodelist if n in pos]
        nx.draw_networkx_nodes(G, pos, nodelist=safe_nodes, node_color=color,
                               node_shape=shape, node_size=size, alpha=alpha, ax=ax)

    # mapping_concepts = [n for n, d in G.nodes(data=True) if d['type'] == 'mapping']
    datasets = [n for n, d in G.nodes(data=True) if d['type'] == 'dataset']
    source_variables = [n for n, d in G.nodes(data=True) if d['type'] == 'svariable']
    target_variables = [n for n, d in G.nodes(data=True) if d['type'] == 'tvariable']
    transformations = [n for n, d in G.nodes(data=True) if d['type'] == 'transformation']
    links = [n for n, d in G.nodes(data=True) if d['type'] == 'link']

    #draw_nodes(mapping_concepts, '#ffcc99', 'o', 2500, 0.8)
    draw_nodes(datasets, 'black', 'o', 5000, 0)
    draw_nodes(source_variables, 'darkblue', 's', 1000, 0)
    draw_nodes(target_variables, '#0047AB', 'o', 1000, 0)
    draw_nodes(transformations, 'coral', 's', 1800, 0)
    draw_nodes(links, 'darkgreen', 'd', 1800, 0)

    for node in source_variables + target_variables:
        if node in pos:
            x, y = pos[node]
            ax.add_patch(patches.Rectangle((x - 0.35, y - 0.3), 0.7, 0.7, linewidth=1, edgecolor='black', facecolor='#66B2E4', alpha=0.8))
    for node in transformations:
        if node in pos:
            x, y = pos[node]
            ax.add_patch(patches.Rectangle((x - 0.25, y - 0.3), 0.5, 0.7, linewidth=1, edgecolor='black', facecolor='#2FAC66', alpha=0.8))
    for node in datasets:
        if node in pos:
            x, y = pos[node]
            ax.add_patch(patches.Rectangle((x - 0.35, y - 0.3), 0.7, 0.7, linewidth=1, edgecolor='black', facecolor="#3B4640", alpha=0.8))
    for node in links:
        if node in pos:
            x, y = pos[node]
            ax.add_patch(patches.Ellipse((x, y), 0.8, 0.7, linewidth=1, edgecolor='black', facecolor="#5F2FAC", alpha=0.8))

    node_labels = {n: G.nodes[n].get('label', str(n)) for n in G.nodes}
    font_color = 'white'
    safe_labels = {n: lbl for n, lbl in node_labels.items() if n in pos and not pd.isna(n)}
    nx.draw_networkx_labels(G, pos, labels=safe_labels, font_size=6, font_weight='bold', font_color=font_color, ax=ax)

    fig.canvas.mpl_connect('button_press_event', lambda event: on_click(event, G, pos))
    plt.title(f"{mapping_concept_id}")
    plt.axis('off')
    plt.show()


start_txt="Lineage Graph"
mapping_file = 'LineageExcel/Mapping_MIMIC_DM_v05_lim.xlsx'
mapping_df, variables_df, datasets_df, link_df = load_data(mapping_file)
#conceptName="Map DM"
G, levels = build_main_graph(mapping_df, variables_df, start_txt)
draw_graph_with_buttons(G, levels, start_txt, mapping_df, variables_df, link_df)






